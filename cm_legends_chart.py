from openpyxl import load_workbook
from bokeh.plotting import figure, ColumnDataSource
from bokeh.models import LabelSet, HoverTool, SingleIntervalTicker
from bokeh.io import output_file, show, curdoc
from bokeh.transform import factor_cmap, factor_mark

def tuple_to_list(x):
    return [x[y].value for y in range(len(x))]



wb = load_workbook('cm_stats.xlsx')
ws = wb.worksheets[0]

# player data
nme = ws['A']
gpn = ws['B']
apn = ws['C']
psn = ws['D']

# data lists
nme_list = tuple_to_list(nme)
gpn_list = tuple_to_list(gpn)
apn_list = tuple_to_list(apn)
psn_list = tuple_to_list(psn)


# data checks
print('names:', nme_list)
print('goals/90:', gpn_list)
print('assists/90:', apn_list)
print(('positions:', psn_list))


data = {
    'goals/90' : gpn_list,
    'assists/90' : apn_list,
    'names' : nme_list,
    'positions' : psn_list
}

source = ColumnDataSource(data)
hover = HoverTool(tooltips=[("index", "index"),("(x,y)", "(@x, @y)")])

positions = sorted(set(psn_list))
p_markers = ['circle_x', 'hex', 'triangle']

n = figure(title="G/90 vs A/90 of Attacking Career Mode Players", width = 1000, height = 620, 
background_fill_color="#fafafa")

n.xaxis.axis_label = "Goals per 90"
n.xaxis.ticker = SingleIntervalTicker(interval=0.1)
n.yaxis.axis_label = "Assists per 90"
n.yaxis.ticker = SingleIntervalTicker(interval=0.1)

n.scatter(x = 'goals/90', y = 'assists/90', source = source, legend_group = "positions", size = 10, 
marker=factor_mark('positions', p_markers, positions), color = factor_cmap('positions', 'Dark2_3', positions),
fill_alpha = 0.5)

labels = LabelSet(x = 'goals/90', y = 'assists/90', text = 'names',  source = source, x_offset = 5, 
y_offset = 5)

n.add_layout(labels)
n.legend.location = "top_right"
n.legend.title = "Positions"

output_file('index.html')
curdoc().theme = 'light_minimal'
show(n)
